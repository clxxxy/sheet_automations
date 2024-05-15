import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import smtplib
import ssl
import mimetypes
from email.message import EmailMessage

#importando dados
data = pd.read_excel('data/VendaCarros.xlsx')

#selecionando colunas para a tabela pivot
data2 = data[["Fabricante", "ValorVenda", "Ano"]]
data2.head()

#criando tabela pivot
pivot_data = data2.pivot_table(index = "Ano",
                              columns = "Fabricante",
                              values = "ValorVenda",
                              aggfunc = "sum") #soma do valor de venda de cada marca/ano

#exportando tabela pivot
pivot_data.to_excel("data/pivot_data.xlsx", "Relatorio")

#ler pasta de trabalho e planilha
wb = load_workbook("data/pivot_data.xlsx")
sheet = wb["Relatorio"]

#referencias das linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#adicionando dados e categorias no gráfico
barchart = BarChart()

data = Reference(sheet,
                 min_col = min_column+1,
                 max_col = max_column,
                 min_row = min_row,
                 max_row = max_row)

categories = Reference(sheet,
                       min_col = min_column,
                       max_col = min_column,
                       min_row = min_row+1,
                       max_row = max_row)

barchart.add_data(data, titles_from_data = True)
barchart.set_categories(categories)

#criando o gráfico
sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricante"
barchart.style = 2

wb.save("data/barchart.xlsx")

#incluindo formula
# sheet["B6"] = "=SUM(B2:B5)"
# sheet["B6"].style = "Currency"

for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    sheet[f"{letter}{max_row+1}"].style = "Currency"

wb.save('data/VendasPorFabricante.xlsx')